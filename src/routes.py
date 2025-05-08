import io
import pandas as pd
from flask import render_template, redirect, url_for, flash, send_file, abort, Blueprint, request
from sqlalchemy import desc
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, date, time
from flask import Flask

from database import database as db 
from database_model import PatientRecord
from forms import RecordForm




def register_routes(app: Flask):

    @app.route("/")
    @app.route("/dashboard")
    def dashboard():
        try:
            records = PatientRecord.query.order_by(desc(PatientRecord.created_at)).all()
        except Exception as e:
            flash(f"Error fetching records: {e}", "danger")
            records = []
        return render_template("dashboard.html", records=records)



    @app.route("/add", methods=["GET", "POST"])
    def add_record():
        form = RecordForm()
        if form.validate_on_submit():
            try:
                new_record = PatientRecord()
                form.populate_obj(new_record)

                for field_name, field in form._fields.items():
                    if field.type in ("FloatField", "IntegerField"):
                        if field.data is None:
                            setattr(new_record, field_name, None)

                db.session.add(new_record)
                db.session.commit()
                flash("Record added successfully!", "success")
                return redirect(url_for("dashboard"))
            except Exception as e:
                db.session.rollback()
                flash(f"Error adding record: {e}", "danger")
                app.logger.error(f"Error adding record: {e}")
        else:
            if request.method == "POST":
                app.logger.warning(f"Form validation failed: {form.errors}")
                flash("One or more fields are invalid", "danger")

        return render_template(
            "form.html",
            form=form,
            form_action=url_for("add_record"),
            form_title="Add New Record",
        )



    @app.route("/edit/<int:record_id>", methods=["GET", "POST"])
    def edit_record(record_id):
        record = PatientRecord.query.get_or_404(record_id)
        form = RecordForm(obj=record)
        if form.validate_on_submit():
            try:
                form.populate_obj(record)

                for field_name, field in form._fields.items():
                    if field.type in ("FloatField", "IntegerField"):
                        if field.data is None:
                            setattr(record, field_name, None)

                db.session.commit()
                flash("Record updated successfully!", "success")
                return redirect(url_for("dashboard"))
            except Exception as e:
                db.session.rollback()
                flash(f"Error updating record: {e}", "danger")
                app.logger.error(f"Error updating record {record_id}: {e}")
        elif request.method == "POST":
            app.logger.warning(f"Form validation failed for edit {record_id}: {form.errors}")
            flash("One or more fields are invalid", "danger")

        return render_template(
            "form.html",
            form=form,
            form_action=url_for("edit_record", record_id=record_id),
            form_title="Edit Record",
        )



    @app.route("/delete/<int:record_id>", methods=["POST"])
    def delete_record(record_id):
        record = PatientRecord.query.get_or_404(record_id)
        try:
            db.session.delete(record)
            db.session.commit()
            flash("Record deleted successfully!", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error deleting record: {e}", "danger")
            app.logger.error(f"Error deleting record {record_id}: {e}")
        return redirect(url_for("dashboard"))



    @app.route("/print/<int:record_id>")
    def print_record(record_id):
        record = PatientRecord.query.get_or_404(record_id)
        return render_template("print.html", record=record)
        # , format_value=app.jinja_env.filters['fmt']



    @app.route("/download")
    def download_records():
        try:
            records = PatientRecord.query.order_by(desc(PatientRecord.created_at)).all()
            if not records:
                flash("No records to download.", "info")
                return redirect(url_for("dashboard"))

            column_order = [c.name for c in PatientRecord.__table__.columns]

            data = []
            for record in records:
                record_dict = {}
                for col_name in column_order:
                    value = getattr(record, col_name, None)
                    if isinstance(value, (datetime, date, time)):
                        record_dict[col_name] = value.isoformat() if value else None
                    elif isinstance(value, bool):
                        record_dict[col_name] = "Yes" if value else "No"
                    else:
                        record_dict[col_name] = value
                data.append(record_dict)

            data_frame = pd.DataFrame(data, columns=column_order)
            sheet_name = "patient-records"

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                data_frame.to_excel(writer, index=False, sheet_name=sheet_name)
                format_excel_sheet(writer, data_frame, sheet_name, "left")
            output.seek(0)

            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                download_name="patient_records.xlsx",
                as_attachment=True,
            )
        except Exception as e:

            flash(f"Error generating Excel file: {e}", "danger")
            app.logger.error(f"Error generating Excel file: {e}")
            return redirect(url_for("dashboard"))



def format_excel_sheet(writer, dataframe, sheet_name, default_alignment):
    
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    alignment_style = Alignment(horizontal=default_alignment, vertical='top', wrap_text=False)

    for i, col_name in enumerate(dataframe.columns):
        column_letter = get_column_letter(i + 1)  # openpyxl columns are indexed from 1
        max_length = 0

        if dataframe[col_name].name: 
            max_length = max(max_length, len(str(dataframe[col_name].name)))

        if not dataframe[col_name].empty:
            column_data = dataframe[col_name].astype(str)
            max_cell_length = column_data.map(len).max()
            max_length = max(max_length, int(max_cell_length)) 

        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment_style